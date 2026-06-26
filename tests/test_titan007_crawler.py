import tempfile
import sys
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
import requests

from main import (
    DATA_SHEET_NAME,
    DEFAULT_TEMPLATE,
    METADATA_SHEET,
    CurrentOdds,
    MatchInfo,
    OddsRecord,
    ScheduleSnapshot,
    changed_schedule_ids,
    combine_export_rows,
    combine_export_records,
    crawl_complete_company_schedule_records,
    crawl_historical_date_records,
    ensure_aomen_company_id,
    fetch_match_histories,
    fetch_text,
    historical_output_path,
    historical_week_output_path,
    is_aomen_change_detail_response,
    is_aomen_change_detail_url,
    filter_matches_by_date_window,
    filter_today_matches,
    parse_historical_results_page,
    parse_change_detail_history,
    parse_bfdata,
    parse_goal_xml,
    parse_odds_history,
    read_selection_matches_from_workbook,
    record_signatures_by_schedule,
    resolve_runtime_path,
    sync_workbook,
    write_workbook,
)


def make_bfdata_line(schedule_id="2986245", state="0", home_rank="11", away_rank="7"):
    fields = [""] * 69
    fields[0] = schedule_id
    fields[1] = "#9DA8A0"
    fields[2] = "印度超"
    fields[5] = "东北联"
    fields[8] = "莫哈末丹"
    fields[11] = "19:30"
    fields[12] = "2026,4,19,19,30,00"
    fields[13] = state
    fields[14] = "0"
    fields[15] = "1"
    fields[16] = "0"
    fields[17] = "1"
    fields[22] = home_rank
    fields[23] = away_rank
    fields[29] = "1.5"
    fields[36] = "5-19"
    fields[43] = "2026"
    fields[46] = "3.25"
    return f'A[1]="{"^".join(fields)}".split("^");'


class Titan007CrawlerTests(unittest.TestCase):
    def test_parse_historical_results_page_extracts_finished_matches(self):
        html = """
        <table id='table_live'>
          <tr height=18 align=center id='tr1_1' sId='2990614'>
            <td><span>中冠</span><span></span></td>
            <td>26日13:00</td>
            <td>完</td>
            <td align=right><span name='order'><font>[12]</font></span>沈阳满天星奥美</td>
            <td><font>2</font>-<font>1</font></td>
            <td align=left>本溪清帆物流园<span name='order'><font>[8]</font></span></td>
            <td><font>2</font>-<font>0</font></td>
            <td id='hdp_1' val=''>平手</td>
            <td id='ou_1' val=''>2.5</td>
            <td></td>
          </tr>
        </table>
        """

        snapshot = parse_historical_results_page(html, datetime(2026, 5, 26).date())

        self.assertEqual(snapshot.schedule_ids, ["2990614"])
        match = snapshot.matches["2990614"]
        self.assertEqual(match.league, "中冠")
        self.assertEqual(match.event_time, "26日13:00")
        self.assertEqual(match.status, "完")
        self.assertEqual(match.home_team, "沈阳满天星奥美")
        self.assertEqual(match.away_team, "本溪清帆物流园")
        self.assertEqual(match.home_rank, "12")
        self.assertEqual(match.away_rank, "8")
        self.assertEqual(match.score, "2-1")
        self.assertEqual(match.half_score, "2-0")
        self.assertEqual(match.match_time, datetime(2026, 5, 26, 13, 0))
        self.assertEqual(snapshot.current_odds["2990614"].asian_line, "平手")
        self.assertEqual(snapshot.current_odds["2990614"].total_line, "2.5")

    def test_historical_output_path_uses_requested_date(self):
        self.assertEqual(
            historical_output_path(Path("out"), datetime(2026, 5, 26).date()),
            Path("out") / "titan007_data_20260526.xlsx",
        )

    def test_historical_week_output_path_uses_start_and_end_dates(self):
        self.assertEqual(
            historical_week_output_path(Path("out"), datetime(2026, 5, 26).date()),
            Path("out") / "titan007_data_20260526_20260601.xlsx",
        )

    def test_parse_bfdata_extracts_match_fields_and_formats_pending_scores(self):
        matches = parse_bfdata("var matchcount=1;\r" + make_bfdata_line(state="0"))

        match = matches["2986245"]
        self.assertEqual(match.league, "印度超")
        self.assertEqual(match.event_time, "19日19:30")
        self.assertEqual(match.home_team, "东北联")
        self.assertEqual(match.away_team, "莫哈末丹")
        self.assertEqual(match.home_rank, "11")
        self.assertEqual(match.away_rank, "7")
        self.assertEqual(match.status, "未")
        self.assertEqual(match.score, "-")
        self.assertEqual(match.half_score, "-")
        self.assertEqual(match.total_line, "3/3.5")
        self.assertEqual(match.match_time, datetime(2026, 5, 19, 19, 30))

    def test_parse_bfdata_formats_finished_scores(self):
        matches = parse_bfdata("var matchcount=1;\r" + make_bfdata_line(state="-1"))

        match = matches["2986245"]
        self.assertEqual(match.score, "0-1")
        self.assertEqual(match.half_score, "0-1")
        self.assertEqual(match.status, "完")

    def test_parse_bfdata_normalizes_rank_prefixes_to_numeric_values(self):
        matches = parse_bfdata(
            "var matchcount=1;\r"
            + make_bfdata_line(state="-1", home_rank="巴西甲20", away_rank="澳昆甲2-7")
        )

        match = matches["2986245"]
        self.assertEqual(match.home_rank, "20")
        self.assertEqual(match.away_rank, "7")

    def test_parse_goal_xml_returns_company_schedule_ids_and_current_odds(self):
        xml = """<?xml version='1.0' encoding='UTF-8'?>
        <c>
          <match>
            <m>2986245,17252901,1.75,0.92,0.78,154321528,1.18,5.05,7.60,19810392,3.25,0.80,0.82</m>
          </match>
          <ids>2986245,2944703,</ids>
          <jcIds></jcIds>
          <isMaintain>0</isMaintain>
        </c>
        """

        goal_data = parse_goal_xml(xml)

        self.assertEqual(goal_data.schedule_ids, ["2986245", "2944703"])
        self.assertEqual(goal_data.current_odds["2986245"].asian_line, "球半/两球")
        self.assertEqual(goal_data.current_odds["2986245"].asian_up, 0.92)
        self.assertEqual(goal_data.current_odds["2986245"].asian_down, 0.78)
        self.assertEqual(goal_data.current_odds["2986245"].total_line, "3/3.5")
        self.assertEqual(goal_data.current_odds["2986245"].total_big, 0.8)
        self.assertEqual(goal_data.current_odds["2986245"].total_small, 0.82)

    def test_parse_odds_history_extracts_company_column_and_sorts_oldest_first(self):
        html = """
        <table id="oddsDetail">
          <tr><th>澳*</th><th>Crow*</th><th>比分</th><th>变化时间</th></tr>
          <tr><td>球半/两球<br><span>0.92</span>&nbsp;<span>0.78</span></td><td></td><td></td><td>5-19 13:02</td></tr>
          <tr><td><font color="red">*</font>半球<br><span>0.76</span>&nbsp;<span>0.94</span></td><td></td><td></td><td>5-17 16:35</td></tr>
        </table>
        """

        records = parse_odds_history(
            html,
            kind="asian",
            match_year=2026,
            match_datetime=datetime(2026, 5, 19, 19, 30),
        )

        self.assertEqual(
            records,
            [
                OddsRecord(datetime(2026, 5, 17, 16, 35), 0.76, "受让半球", 0.94),
                OddsRecord(datetime(2026, 5, 19, 13, 2), 0.92, "球半/两球", 0.78),
            ],
        )

    def test_parse_total_history_extracts_big_line_small(self):
        html = """
        <table id="oddsDetail">
          <tr><th>澳*</th><th>比分</th><th>变化时间</th></tr>
          <tr><td>3/3.5<br><span>0.80</span>&nbsp;<span>0.82</span></td><td></td><td>5-19 13:02</td></tr>
        </table>
        """

        records = parse_odds_history(
            html,
            kind="total",
            match_year=2026,
            match_datetime=datetime(2026, 5, 19, 19, 30),
        )

        self.assertEqual(records, [OddsRecord(datetime(2026, 5, 19, 13, 2), 0.8, "3/3.5", 0.82)])

    def test_parse_change_detail_history_keeps_only_status_immediate_records(self):
        html = """
        <table>
          <tr><td>时间</td><td>比分</td><td>圣塔菲</td><td>盘</td><td>普拉腾斯</td><td>变化时间</td><td>状态</td></tr>
          <tr><td>83</td><td>2-1</td><td>0.66</td><td>平手</td><td>1.10</td><td>5-20 09:45</td><td>滚</td></tr>
          <tr><td></td><td></td><td>0.78</td><td>平手/半球</td><td>1.00</td><td>5-20 07:41</td><td>即</td></tr>
          <tr><td></td><td></td><td>封</td><td colspan="3">5-20 07:40</td><td>即</td></tr>
          <tr><td></td><td></td><td>0.84</td><td>半球</td><td>0.94</td><td>5-19 13:41</td><td>即</td></tr>
        </table>
        """

        records = parse_change_detail_history(
            html,
            kind="asian",
            match_year=2026,
            match_datetime=datetime(2026, 5, 20, 8, 0),
        )

        self.assertEqual(
            records,
            [
                OddsRecord(datetime(2026, 5, 19, 13, 41), 0.84, "半球", 0.94, "即"),
                OddsRecord(datetime(2026, 5, 20, 7, 41), 0.78, "平手/半球", 1.0, "即"),
            ],
        )

    def test_parse_change_detail_total_history_extracts_immediate_big_line_small(self):
        html = """
        <table>
          <tr><td>时间</td><td>比分</td><td>大球</td><td>进球数</td><td>小球</td><td>变化时间</td><td>状态</td></tr>
          <tr><td></td><td></td><td>0.80</td><td>2.5/3</td><td>0.82</td><td>5-20 07:41</td><td>即</td></tr>
          <tr><td>10</td><td>0-0</td><td>0.70</td><td>2.5</td><td>0.92</td><td>5-20 08:10</td><td>滚</td></tr>
        </table>
        """

        records = parse_change_detail_history(
            html,
            kind="total",
            match_year=2026,
            match_datetime=datetime(2026, 5, 20, 8, 0),
        )

        self.assertEqual(records, [OddsRecord(datetime(2026, 5, 20, 7, 41), 0.8, "2.5/3", 0.82, "即")])

    def test_combine_export_rows_keeps_base_info_on_first_row_and_adds_separator(self):
        matches = parse_bfdata("var matchcount=1;\r" + make_bfdata_line(state="-1"))
        asian = {
            "2986245": [
                OddsRecord(datetime(2026, 5, 17, 16, 35), 0.76, "受让半球", 0.94),
                OddsRecord(datetime(2026, 5, 19, 13, 2), 0.92, "球半/两球", 0.78),
                OddsRecord(datetime(2026, 5, 19, 14, 2), 0.95, "球半", 0.75),
            ]
        }
        total = {
            "2986245": [
                OddsRecord(datetime(2026, 5, 19, 13, 2), 0.8, "3/3.5", 0.82),
                OddsRecord(datetime(2026, 5, 19, 14, 2), 0.85, "3", 0.77),
            ]
        }

        rows = combine_export_rows(
            matches,
            ["2986245"],
            asian,
            total,
            current_odds={
                "2986245": CurrentOdds("球半", 0.91, 0.89, "3/3.5", 0.8, 0.82),
            },
        )

        self.assertEqual(len(rows), 4)
        self.assertEqual(rows[0][:9], ["印度超", "19日19:30", "完", "[11] 东北联", "0-1", "[7] 莫哈末丹", 0.91, "球半", 0.89])
        self.assertEqual(rows[0][9:21], [datetime(2026, 5, 17, 16, 35), 0.76, "受让半球", 0.94, 2, 2, datetime(2026, 5, 19, 13, 2), 0.8, "3/3.5", 0.82, 1, 1])
        self.assertEqual(rows[1][:9], [None] * 9)
        self.assertEqual(rows[1][9:21], [datetime(2026, 5, 19, 13, 2), 0.92, "球半/两球", 0.78, "", "", datetime(2026, 5, 19, 14, 2), 0.85, "3", 0.77, "", ""])
        self.assertEqual(rows[2][9:15], [datetime(2026, 5, 19, 14, 2), 0.95, "球半", 0.75, "", ""])
        self.assertEqual(rows[3], [None] * 21)

    def test_asian_handicap_change_counts_adjacent_line_changes(self):
        match = MatchInfo("1001", "League", "5-23 19:30", "Home", "-", "Away", "-", "2.5", datetime(2026, 5, 23, 19, 30))
        rows = combine_export_rows(
            {"1001": match},
            ["1001"],
            {
                "1001": [
                    OddsRecord(datetime(2026, 5, 23, 12, 0), 0.82, "A", 1.0),
                    OddsRecord(datetime(2026, 5, 23, 13, 0), 0.84, "A", 0.98),
                    OddsRecord(datetime(2026, 5, 23, 14, 0), 0.86, "B", 0.96),
                    OddsRecord(datetime(2026, 5, 23, 15, 0), 0.88, "A", 0.94),
                ]
            },
            {},
        )

        self.assertEqual(rows[0][13:15], [3, 2])

    def test_total_handicap_change_counts_adjacent_line_changes(self):
        match = MatchInfo("1001", "League", "5-23 19:30", "Home", "-", "Away", "-", "2.5", datetime(2026, 5, 23, 19, 30))
        rows = combine_export_rows(
            {"1001": match},
            ["1001"],
            {},
            {
                "1001": [
                    OddsRecord(datetime(2026, 5, 23, 12, 0), 0.82, "2.5", 1.0),
                    OddsRecord(datetime(2026, 5, 23, 13, 0), 0.84, "2.5", 0.98),
                    OddsRecord(datetime(2026, 5, 23, 14, 0), 0.86, "3", 0.96),
                    OddsRecord(datetime(2026, 5, 23, 15, 0), 0.88, "2.5", 0.94),
                ]
            },
        )

        self.assertEqual(rows[0][19:21], [3, 2])

    def test_write_workbook_creates_reference_columns(self):
        rows = [["印度超", "19日19:30", "未", "[11] 东北联", "-", "[7] 莫哈末丹", 0.91, "球半", 0.89, None, None, None, None, None, None, None, None, None, None, None, None]]

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "out.xlsx"
            write_workbook(rows, path, template_path=None)
            workbook = load_workbook(path)
            sheet = workbook[DATA_SHEET_NAME]

            self.assertEqual(sheet.max_column, 21)
            self.assertEqual([sheet.cell(1, c).value for c in range(1, 22)], [
                "联赛",
                "时间",
                "状态",
                "比赛球队",
                "比分",
                "比赛球队",
                "指数",
                "盘口",
                "指数",
                "时间节点",
                "上盘",
                "盘",
                "下盘",
                "赔率变动",
                "盘口变动",
                "时间节点",
                "大球",
                "盘口",
                "小球",
                "赔率变动",
                "盘口变动",
            ])
            self.assertEqual(sheet.cell(2, 1).value, "印度超")

    def test_write_workbook_colors_only_handicap_change_count_cells(self):
        rows = [["印度超", "19日19:30", "未", "[11] 东北联", "-", "[7] 莫哈末丹", 0.91, "球半", 0.89, None, None, None, None, 2, 3, None, None, None, None, 1, 4]]

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "out.xlsx"
            write_workbook(rows, path, template_path=None)
            sheet = load_workbook(path)[DATA_SHEET_NAME]

            self.assertTrue(sheet.cell(2, 15).font.color.rgb.endswith("1E3A8A"))
            self.assertTrue(sheet.cell(2, 21).font.color.rgb.endswith("FF0000"))
            self.assertNotEqual(sheet.cell(2, 14).font.color.type, "rgb")
            self.assertNotEqual(sheet.cell(2, 20).font.color.type, "rgb")

    def test_default_template_uses_league_match_data_workbook(self):
        self.assertEqual(DEFAULT_TEMPLATE, "联赛比赛数据.xlsx")

    def test_sync_workbook_migrates_legacy_sheet_to_new_template_name(self):
        records = make_export_records(schedule_id="2986245")

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "sync.xlsx"
            sync_workbook(records, path, template_path=None)
            workbook = load_workbook(path)
            workbook[DATA_SHEET_NAME].title = "比赛结果"
            workbook.save(path)
            workbook.close()

            sync_workbook(records, path, template_path=DEFAULT_TEMPLATE)
            workbook = load_workbook(path, data_only=True)

            self.assertIn(DATA_SHEET_NAME, workbook.sheetnames)
            self.assertNotIn("比赛结果", workbook.sheetnames)
            self.assertEqual(workbook[DATA_SHEET_NAME].column_dimensions["A"].width, 13.0)

    def test_resolve_runtime_path_uses_pyinstaller_bundle_for_relative_files(self):
        original_meipass = getattr(sys, "_MEIPASS", None)
        had_meipass = hasattr(sys, "_MEIPASS")
        with tempfile.TemporaryDirectory() as tmp:
            bundled_file = Path(tmp) / "template.xlsx"
            bundled_file.write_text("template", encoding="utf-8")
            sys._MEIPASS = tmp
            try:
                self.assertEqual(resolve_runtime_path("template.xlsx"), bundled_file)
            finally:
                if had_meipass:
                    sys._MEIPASS = original_meipass
                else:
                    delattr(sys, "_MEIPASS")

    def test_sync_workbook_writes_hidden_metadata_sheet(self):
        records = make_export_records(schedule_id="2986245")

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "sync.xlsx"
            sync_workbook(records, path, template_path=None)
            workbook = load_workbook(path, data_only=True)

            self.assertIn(METADATA_SHEET, workbook.sheetnames)
            self.assertEqual(workbook[METADATA_SHEET].sheet_state, "hidden")
            self.assertEqual(workbook[METADATA_SHEET].cell(1, 1).value, "version")

    def test_sync_workbook_twice_does_not_duplicate_existing_match(self):
        records = make_export_records(schedule_id="2986245")

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "sync.xlsx"
            sync_workbook(records, path, template_path=None)
            sync_workbook(records, path, template_path=None)
            sheet = load_workbook(path, data_only=True)[DATA_SHEET_NAME]

            base_rows = [row for row in range(2, sheet.max_row + 1) if sheet.cell(row, 1).value]
            self.assertEqual(len(base_rows), 1)

    def test_sync_workbook_updates_existing_match_and_odds_values(self):
        original = make_export_records(schedule_id="2986245", state="0", asian_up=0.76, asian_down=0.94)
        updated = make_export_records(schedule_id="2986245", state="-1", asian_up=0.8, asian_down=0.9)

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "sync.xlsx"
            sync_workbook(original, path, template_path=None)
            sync_workbook(updated, path, template_path=None)
            sheet = load_workbook(path, data_only=True)[DATA_SHEET_NAME]

            self.assertEqual(sheet.cell(2, 5).value, "0-1")
            self.assertEqual(sheet.cell(2, 11).value, 0.8)
            self.assertEqual(sheet.cell(2, 13).value, 0.9)

    def test_sync_workbook_preserves_existing_values_when_latest_crawl_is_partial(self):
        original = make_export_records(schedule_id="2986245", state="0", asian_up=0.76, asian_down=0.94)
        matches = parse_bfdata("var matchcount=1;\r" + make_bfdata_line(schedule_id="2986245", state="-1"))
        partial = combine_export_records(matches, ["2986245"], {}, {})

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "sync.xlsx"
            sync_workbook(original, path, template_path=None)
            sync_workbook(partial, path, template_path=None)
            sheet = load_workbook(path, data_only=True)[DATA_SHEET_NAME]

            self.assertEqual(sheet.cell(2, 5).value, "0-1")
            self.assertEqual(sheet.cell(2, 11).value, 0.76)
            self.assertEqual(sheet.cell(2, 12).value, "受让半球")
            self.assertEqual(sheet.cell(2, 13).value, 0.94)
            self.assertEqual(sheet.cell(2, 17).value, 0.91)
            self.assertEqual(sheet.cell(2, 18).value, "3/3.5")

    def test_sync_workbook_preserves_old_match_missing_from_latest_crawl(self):
        first_run = make_export_records(schedule_id="2986245") + make_export_records(schedule_id="2944703")
        second_run = make_export_records(schedule_id="2986245", state="-1")

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "sync.xlsx"
            sync_workbook(first_run, path, template_path=None)
            sync_workbook(second_run, path, template_path=None)
            sheet = load_workbook(path, data_only=True)[DATA_SHEET_NAME]

            base_ids = []
            metadata = load_workbook(path, data_only=True)[METADATA_SHEET]
            headers = [metadata.cell(1, c).value for c in range(1, metadata.max_column + 1)]
            schedule_col = headers.index("schedule_id") + 1
            row_type_col = headers.index("row_type") + 1
            for row in range(2, metadata.max_row + 1):
                if metadata.cell(row, row_type_col).value == "match":
                    base_ids.append(metadata.cell(row, schedule_col).value)

            self.assertEqual(base_ids, ["2986245", "2944703"])
            self.assertEqual(sum(1 for row in range(2, sheet.max_row + 1) if sheet.cell(row, 1).value), 2)

    def test_fetch_match_histories_retries_transient_odd_page_failure(self):
        match = MatchInfo("2986245", "印度超", "19日19:30", "东北联", "-", "莫哈末丹", "-", "3/3.5", datetime(2026, 5, 19, 19, 30))
        session = FakeSession(
            {
                "changeDetail/handicap.aspx?id=2986245&companyID=1": [
                    FakeResponse(443, ""),
                    FakeResponse(200, change_detail_html("球半/两球", "0.92", "0.78")),
                ],
                "changeDetail/overunder.aspx?id=2986245&companyID=1": [FakeResponse(200, change_detail_html("3/3.5", "0.80", "0.82", line_header="进球数"))],
            }
        )

        asian, total = fetch_match_histories(session, "2986245", match, retries=1, retry_delay=0, log_failures=False)

        self.assertEqual(asian, [OddsRecord(datetime(2026, 5, 19, 13, 2), 0.92, "球半/两球", 0.78, "即")])
        self.assertEqual(total, [OddsRecord(datetime(2026, 5, 19, 13, 2), 0.8, "3/3.5", 0.82, "即")])

    def test_fetch_match_histories_rejects_non_aomen_change_detail_response(self):
        match = MatchInfo("2986245", "印度超", "19日19:30", "东北联", "-", "莫哈末丹", "-", "3/3.5", datetime(2026, 5, 19, 19, 30))
        session = FakeSession(
            {
                "changeDetail/handicap.aspx?id=2986245&companyID=1": [
                    FakeResponse(200, change_detail_html("球半/两球", "0.92", "0.78").replace("company=澳*", "company=Crow*"))
                ],
                "changeDetail/overunder.aspx?id=2986245&companyID=1": [
                    FakeResponse(200, change_detail_html("3/3.5", "0.80", "0.82", line_header="进球数").replace("company=澳*", "company=Crow*"))
                ],
            }
        )

        asian, total = fetch_match_histories(session, "2986245", match, retries=0, retry_delay=0, log_failures=False)

        self.assertEqual(asian, [])
        self.assertEqual(total, [])

    def test_change_detail_company_validation_requires_aomen(self):
        self.assertTrue(is_aomen_change_detail_url("https://vip.titan007.com/changeDetail/handicap.aspx?id=1&companyID=1&l=0"))
        self.assertFalse(is_aomen_change_detail_url("https://vip.titan007.com/changeDetail/handicap.aspx?id=1&companyID=3&l=0"))
        self.assertTrue(is_aomen_change_detail_response(change_detail_html("球半", "0.8", "0.9")))
        self.assertFalse(is_aomen_change_detail_response(change_detail_html("球半", "0.8", "0.9").replace("company=澳*", "company=Crow*")))
        self.assertTrue(is_aomen_change_detail_response('<a href="handicap.aspx?id=1&companyid=1&l=0">亚让</a><iframe src="chartFlash.aspx?company=&scheid=1"></iframe>'))

    def test_company_id_guard_rejects_non_aomen_company(self):
        with self.assertRaises(ValueError):
            ensure_aomen_company_id(3)

    def test_fetch_match_histories_keeps_total_when_asian_page_exhausts_retries(self):
        match = MatchInfo("2986245", "印度超", "19日19:30", "东北联", "-", "莫哈末丹", "-", "3/3.5", datetime(2026, 5, 19, 19, 30))
        session = FakeSession(
            {
                "changeDetail/handicap.aspx?id=2986245&companyID=1": [FakeResponse(443, ""), FakeResponse(443, "")],
                "changeDetail/overunder.aspx?id=2986245&companyID=1": [FakeResponse(200, change_detail_html("3/3.5", "0.80", "0.82", line_header="进球数"))],
            }
        )

        asian, total = fetch_match_histories(session, "2986245", match, retries=1, retry_delay=0, log_failures=False)

        self.assertEqual(asian, [])
        self.assertEqual(total, [OddsRecord(datetime(2026, 5, 19, 13, 2), 0.8, "3/3.5", 0.82, "即")])

    def test_fetch_text_decodes_gb2312_change_detail_pages(self):
        session = FakeSession(
            {
                "changeDetail": [FakeResponse(200, "\u72b6\u6001", encoding="gb2312", apparent_encoding="GB2312")],
            }
        )

        self.assertEqual(fetch_text(session, "https://vip.titan007.com/changeDetail/handicap.aspx?id=1", referer="x"), "\u72b6\u6001")

    def test_filter_today_matches_keeps_company_order_for_current_date(self):
        today_match = MatchInfo("2986245", "印度超", "19日19:30", "东北联", "-", "莫哈末丹", "-", "3/3.5", datetime(2026, 5, 19, 19, 30))
        tomorrow_match = MatchInfo("2944703", "巴西甲", "20日06:00", "主队", "-", "客队", "-", "2.5", datetime(2026, 5, 20, 6, 0))

        matches = filter_today_matches(
            {"2986245": today_match, "2944703": tomorrow_match},
            ["2944703", "2986245"],
            today=datetime(2026, 5, 19).date(),
        )

        self.assertEqual(matches, [today_match])

    def test_filter_matches_by_date_window_keeps_company_order_and_inclusive_bounds(self):
        before_edge = MatchInfo("1001", "A", "5-16 19:30", "H1", "-", "A1", "-", "2.5", datetime(2026, 5, 16, 19, 30))
        center = MatchInfo("1002", "B", "5-23 20:30", "H2", "-", "A2", "-", "2.5", datetime(2026, 5, 23, 20, 30))
        after_edge = MatchInfo("1003", "C", "5-30 21:30", "H3", "-", "A3", "-", "2.5", datetime(2026, 5, 30, 21, 30))
        too_old = MatchInfo("1004", "D", "5-15 18:00", "H4", "-", "A4", "-", "2.5", datetime(2026, 5, 15, 18, 0))
        too_new = MatchInfo("1005", "E", "5-31 18:00", "H5", "-", "A5", "-", "2.5", datetime(2026, 5, 31, 18, 0))

        matches = filter_matches_by_date_window(
            {
                "1001": before_edge,
                "1002": center,
                "1003": after_edge,
                "1004": too_old,
                "1005": too_new,
            },
            ["1005", "missing", "1003", "1001", "1004", "1002"],
            center_date=datetime(2026, 5, 23).date(),
        )

        self.assertEqual(matches, [after_edge, before_edge, center])

    def test_read_selection_matches_from_workbook_returns_only_window_matches(self):
        in_range = MatchInfo("1001", "League", "5-20 19:30", "Home", "-", "Away", "-", "2.5", datetime(2026, 5, 20, 19, 30))
        out_of_range = MatchInfo("1002", "League", "5-31 19:30", "OldHome", "-", "OldAway", "-", "2.5", datetime(2026, 5, 31, 19, 30))
        records = combine_export_records(
            {"1001": in_range, "1002": out_of_range},
            ["1001", "1002"],
            {},
            {},
        )

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "selection.xlsx"
            sync_workbook(records, path, template_path=None)

            matches = read_selection_matches_from_workbook(
                path,
                center_date=datetime(2026, 5, 23).date(),
            )

        self.assertEqual([match.schedule_id for match in matches], ["1001"])
        self.assertEqual(matches[0].match_time, datetime(2026, 5, 20, 19, 30))
        self.assertEqual(matches[0].home_team, "Home")

    def test_crawl_records_can_limit_to_selected_schedule_ids(self):
        selected_match = MatchInfo("2944703", "巴西甲", "20日06:00", "主队", "-", "客队", "-", "2.5", datetime(2026, 5, 20, 6, 0))
        skipped_match = MatchInfo("2986245", "印度超", "19日19:30", "东北联", "-", "莫哈末丹", "-", "3/3.5", datetime(2026, 5, 19, 19, 30))
        fetched = []

        def fake_fetch(session, schedule_id, match):
            fetched.append(schedule_id)
            return [], []

        records = crawl_complete_company_schedule_records(
            selected_schedule_ids=["2944703"],
            schedule_snapshot=ScheduleSnapshot(
                matches={"2986245": skipped_match, "2944703": selected_match},
                schedule_ids=["2986245", "2944703"],
            ),
            fetch_histories=fake_fetch,
            log_to_console=False,
        )

        self.assertEqual(fetched, ["2944703"])
        self.assertEqual({record.schedule_id for record in records}, {"2944703"})

    def test_historical_crawl_exports_only_matches_with_aomen_odds_records(self):
        first = MatchInfo("1001", "A", "26日13:00", "H1", "1-0", "A1", "0-0", "", datetime(2026, 5, 26, 13, 0))
        second = MatchInfo("1002", "B", "26日15:00", "H2", "2-0", "A2", "1-0", "", datetime(2026, 5, 26, 15, 0))
        snapshot = ScheduleSnapshot(
            matches={"1001": first, "1002": second},
            schedule_ids=["1001", "1002"],
        )
        fetched = []

        def fake_fetch(session, schedule_id, match):
            fetched.append(schedule_id)
            if schedule_id == "1001":
                return [OddsRecord(datetime(2026, 5, 26, 12, 0), 0.8, "平手", 0.9, "即")], []
            return [], []

        records = crawl_historical_date_records(
            datetime(2026, 5, 26).date(),
            schedule_snapshot=snapshot,
            fetch_histories=fake_fetch,
            log_to_console=False,
        )

        self.assertEqual(fetched, ["1001", "1002"])
        self.assertEqual({record.schedule_id for record in records}, {"1001"})

    def test_changed_schedule_ids_detects_score_and_odds_updates(self):
        baseline = make_export_records(schedule_id="2986245", state="0", asian_up=0.76, asian_down=0.94)
        baseline += make_export_records(schedule_id="2944703", state="0", asian_up=0.76, asian_down=0.94)
        score_updated = make_export_records(schedule_id="2986245", state="-1", asian_up=0.76, asian_down=0.94)
        odds_updated = make_export_records(schedule_id="2944703", state="0", asian_up=0.8, asian_down=0.9)

        baseline_signatures = record_signatures_by_schedule(baseline)

        self.assertEqual(changed_schedule_ids(baseline_signatures, score_updated + odds_updated), ["2986245", "2944703"])


if __name__ == "__main__":
    unittest.main()


def odds_history_html(line, first, second):
    return f"""
    <table id="oddsDetail">
      <tr><th>澳*</th><th>比分</th><th>变化时间</th></tr>
      <tr><td>{line}<br><span>{first}</span>&nbsp;<span>{second}</span></td><td></td><td>5-19 13:02</td></tr>
    </table>
    """


def change_detail_html(line, first, second, status="即", line_header="盘"):
    return f"""
    <table>
      <tr><td>时间</td><td>比分</td><td>主队</td><td>{line_header}</td><td>客队</td><td>变化时间</td><td>状态</td></tr>
      <tr><td></td><td></td><td>{first}</td><td>{line}</td><td>{second}</td><td>5-19 13:02</td><td>{status}</td></tr>
      <tr><td colspan="7">company=澳*</td></tr>
    </table>
    """


class FakeResponse:
    def __init__(self, status_code, content, encoding="utf-8", apparent_encoding=None):
        self.status_code = status_code
        self.encoding = encoding
        self.apparent_encoding = apparent_encoding
        self.content = content.encode(encoding)

    def raise_for_status(self):
        if self.status_code >= 400:
            raise requests.HTTPError(f"{self.status_code} Client Error")


class FakeSession:
    def __init__(self, responses_by_url_part):
        self.responses_by_url_part = responses_by_url_part

    def get(self, url, headers=None, timeout=None):
        for key, responses in self.responses_by_url_part.items():
            if key in url:
                if responses:
                    return responses.pop(0)
                return FakeResponse(500, "")
        return FakeResponse(404, "")


def make_export_records(schedule_id="2986245", state="0", asian_up=0.76, asian_down=0.94):
    matches = parse_bfdata("var matchcount=1;\r" + make_bfdata_line(schedule_id=schedule_id, state=state))
    asian = {schedule_id: [OddsRecord(datetime(2026, 5, 17, 16, 35), asian_up, "受让半球", asian_down)]}
    total = {schedule_id: [OddsRecord(datetime(2026, 5, 17, 16, 35), 0.91, "3/3.5", 0.71)]}
    return combine_export_records(matches, [schedule_id], asian, total)
