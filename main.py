from __future__ import annotations

import argparse
import copy
import re
import sys
import time
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field, replace
from datetime import date, datetime, timedelta
from html import unescape
from pathlib import Path
from typing import Iterable
from urllib.parse import parse_qs, unquote, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
# from startup_guard import StartupBlocked, enforce_startup_time_limit


ENTRY_URL = "https://live.titan007.com/index2in1.aspx?id=1"
STATIC_BASE_URL = "https://livestatic.titan007.com"
VIP_BASE_URL = "https://vip.titan007.com"
HISTORICAL_RESULTS_BASE_URL = "https://bf.titan007.com/football"
DEFAULT_TEMPLATE = "联赛比赛数据.xlsx"
DATA_SHEET_NAME = "Sheet1"
LEGACY_DATA_SHEET_NAME = "比赛结果"
AOMEN_COMPANY_ID = 1
AOMEN_COMPANY_NAME = "澳*"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
}
HEADERS_ROW = [
    "联赛",
    "时间",
    "状态",
    "比赛球队",
    "比分",
    "比赛球队",
    "指数",
    "盘口",
    "指数",
    "时间节点",
    "上盘",
    "盘",
    "下盘",
    "赔率变动",
    "盘口变动",
    "时间节点",
    "大球",
    "盘口",
    "小球",
    "赔率变动",
    "盘口变动",
]
METADATA_SHEET = "_metadata"
METADATA_VERSION = "1"
METADATA_HEADERS = [
    "version",
    "visible_row",
    "record_key",
    "schedule_id",
    "row_type",
    "match_order",
    "row_order",
    "asian_key",
    "total_key",
]
SYNC_DEFAULT_OUTPUT = "titan007_data.xlsx"
ASIAN_GOAL_NAMES = [
    "平手",
    "平手/半球",
    "半球",
    "半球/一球",
    "一球",
    "一球/球半",
    "球半",
    "球半/两球",
    "两球",
    "两球/两球半",
    "两球半",
    "两球半/三球",
    "三球",
    "三球/三球半",
    "三球半",
    "三球半/四球",
    "四球",
    "四球/四球半",
    "四球半",
    "四球半/五球",
    "五球",
    "五球/五球半",
    "五球半",
    "五球半/六球",
    "六球",
    "六球/六球半",
    "六球半",
    "六球半/七球",
    "七球",
    "七球/七球半",
    "七球半",
    "七球半/八球",
    "八球",
    "八球/八球半",
    "八球半",
    "八球半/九球",
    "九球",
    "九球/九球半",
    "九球半",
    "九球半/十球",
    "十球",
]


@dataclass(frozen=True)
class MatchInfo:
    schedule_id: str
    league: str
    event_time: str
    home_team: str
    score: str
    away_team: str
    half_score: str
    total_line: str
    match_time: datetime
    status: str = ""
    home_rank: str = ""
    away_rank: str = ""


@dataclass(frozen=True)
class CurrentOdds:
    asian_line: str
    asian_up: float | int | None
    asian_down: float | int | None
    total_line: str
    total_big: float | int | None
    total_small: float | int | None


@dataclass(frozen=True)
class GoalData:
    schedule_ids: list[str]
    current_odds: dict[str, CurrentOdds]


@dataclass(frozen=True)
class ScheduleSnapshot:
    matches: dict[str, MatchInfo]
    schedule_ids: list[str]
    current_odds: dict[str, CurrentOdds] = field(default_factory=dict)


@dataclass(frozen=True)
class OddsRecord:
    time_node: datetime
    first: float | int
    line: str
    second: float | int
    status: str = ""


@dataclass(frozen=True)
class ExportRecord:
    schedule_id: str
    record_key: str
    row_type: str
    match_order: int
    row_order: int
    row_values: tuple[object | None, ...]
    asian_key: str = ""
    total_key: str = ""


@dataclass(frozen=True)
class CrawlEvent:
    type: str
    total: int | None = None
    completed: int | None = None
    schedule_id: str | None = None
    rows: int | None = None
    message: str | None = None


class CrawlCancelled(Exception):
    """Raised when a GUI/user cancellation request stops the crawl."""


def parse_bfdata(text: str) -> dict[str, MatchInfo]:
    matches: dict[str, MatchInfo] = {}
    for raw_fields in re.findall(r'A\[\d+\]\s*=\s*"([^"]*)"\.split\([\'"]\^[\'"]\)', text):
        fields = raw_fields.split("^")
        if len(fields) < 47 or not fields[0]:
            continue

        schedule_id = fields[0]
        match_time = parse_match_datetime(fields)
        state = safe_int(fields[13])
        score = format_score(state, fields[14], fields[15])
        half_score = format_score(state, fields[16], fields[17])

        matches[schedule_id] = MatchInfo(
            schedule_id=schedule_id,
            league=fields[2],
            event_time=format_event_time(fields[36], fields[11]),
            home_team=strip_html(fields[5]),
            score=score,
            away_team=strip_html(fields[8]),
            half_score=half_score,
            total_line=total_goal_to_text(fields[46]),
            match_time=match_time,
            status=format_match_status(state),
            home_rank=normalize_rank(fields[22]),
            away_rank=normalize_rank(fields[23]),
        )
    return matches


def parse_historical_results_page(html: str, match_date: date) -> ScheduleSnapshot:
    soup = BeautifulSoup(html, "html.parser")
    matches: dict[str, MatchInfo] = {}
    current_odds: dict[str, CurrentOdds] = {}
    schedule_ids: list[str] = []

    for row in soup.select("tr[id^='tr1_'][sid]"):
        schedule_id = (row.get("sid") or row.get("sId") or "").strip()
        if not schedule_id:
            continue
        cells = row.find_all("td", recursive=False)
        if len(cells) < 9:
            continue

        league = clean_cell_text(cells[0].find("span") or cells[0])
        event_time = clean_cell_text(cells[1])
        status = clean_cell_text(cells[2])
        home_rank, home_team = parse_historical_team_cell(cells[3])
        score = clean_score_text(cells[4])
        away_rank, away_team = parse_historical_team_cell(cells[5])
        half_score = clean_score_text(cells[6])
        asian_line = clean_cell_text(cells[7])
        total_line = clean_cell_text(cells[8])

        match_time = parse_historical_match_time(event_time, match_date)
        matches[schedule_id] = MatchInfo(
            schedule_id=schedule_id,
            league=league,
            event_time=event_time,
            home_team=home_team,
            score=score,
            away_team=away_team,
            half_score=half_score,
            total_line=total_line,
            match_time=match_time,
            status=status,
            home_rank=home_rank,
            away_rank=away_rank,
        )
        current_odds[schedule_id] = CurrentOdds(
            asian_line=asian_line,
            asian_up=None,
            asian_down=None,
            total_line=total_line,
            total_big=None,
            total_small=None,
        )
        schedule_ids.append(schedule_id)

    return ScheduleSnapshot(matches=matches, schedule_ids=schedule_ids, current_odds=current_odds)


def fetch_historical_schedule_snapshot(match_date: date, session: requests.Session | None = None) -> ScheduleSnapshot:
    session = session or requests.Session()
    session.headers.update(HEADERS)
    url = historical_results_url(match_date)
    response = session.get(url, timeout=30)
    response.raise_for_status()
    encoding = getattr(response, "apparent_encoding", None) or getattr(response, "encoding", None) or "gb18030"
    html = response.content.decode(encoding, errors="replace")
    return parse_historical_results_page(html, match_date)


def historical_results_url(match_date: date) -> str:
    return f"{HISTORICAL_RESULTS_BASE_URL}/Over_{match_date:%Y%m%d}.htm"


def historical_output_path(output_dir: str | Path, match_date: date) -> Path:
    return Path(output_dir) / f"titan007_data_{match_date:%Y%m%d}.xlsx"


def historical_week_output_path(output_dir: str | Path, start_date: date) -> Path:
    end_date = start_date + timedelta(days=6)
    return Path(output_dir) / f"titan007_data_{start_date:%Y%m%d}_{end_date:%Y%m%d}.xlsx"


def parse_historical_team_cell(cell) -> tuple[str, str]:
    clone = BeautifulSoup(str(cell), "html.parser")
    rank = ""
    order = clone.find(attrs={"name": "order"})
    if order is not None:
        rank = normalize_rank(order.get_text(" ", strip=True))
        order.decompose()
    for removable in clone.find_all(["img"]):
        removable.decompose()
    for yellow in clone.find_all(attrs={"name": "yellow"}):
        yellow.decompose()
    return rank, clean_cell_text(clone)


def clean_cell_text(node) -> str:
    return re.sub(r"\s+", " ", node.get_text("", strip=True)).strip()


def clean_score_text(node) -> str:
    text = node.get_text("", strip=True)
    text = text.replace("：", "-").replace(":", "-")
    return re.sub(r"\s+", "", text)


def parse_historical_match_time(event_time: str, match_date: date) -> datetime:
    numbers = [int(value) for value in re.findall(r"\d{1,2}", event_time or "")]
    if len(numbers) >= 3:
        _day, hour, minute = numbers[-3], numbers[-2], numbers[-1]
    elif len(numbers) >= 2:
        hour, minute = numbers[-2], numbers[-1]
    else:
        hour, minute = 0, 0
    return datetime(match_date.year, match_date.month, match_date.day, hour, minute)


def parse_goal_xml(xml_text: str | bytes) -> GoalData:
    root = ET.fromstring(xml_text)
    ids_text = root.findtext("ids", default="")
    schedule_ids = [item for item in ids_text.split(",") if item]
    current_odds: dict[str, CurrentOdds] = {}

    for node in root.findall(".//match/m"):
        if not node.text:
            continue
        fields = node.text.split(",")
        if len(fields) < 13:
            continue
        current_odds[fields[0]] = CurrentOdds(
            asian_line=asian_goal_to_text(fields[2]),
            asian_up=to_number(fields[3]),
            asian_down=to_number(fields[4]),
            total_line=total_goal_to_text(fields[10]),
            total_big=to_number(fields[11]),
            total_small=to_number(fields[12]),
        )

    return GoalData(schedule_ids=schedule_ids, current_odds=current_odds)


def fetch_company_schedule_snapshot(company_id: int = 1, session: requests.Session | None = None) -> ScheduleSnapshot:
    ensure_aomen_company_id(company_id)
    session = session or requests.Session()
    session.headers.update(HEADERS)
    session.get(ENTRY_URL, timeout=30).raise_for_status()

    stamp = int(time.time() * 1000)
    bfdata_text = fetch_text(
        session,
        f"{STATIC_BASE_URL}/vbsxml/bfdata_ut.js?r=007{stamp}",
        referer=ENTRY_URL,
    )
    goal_xml = fetch_text(
        session,
        f"{STATIC_BASE_URL}/vbsxml/goal{company_id}.xml?r=007{stamp}",
        referer=ENTRY_URL,
    )

    matches = parse_bfdata(bfdata_text)
    goal_data = parse_goal_xml(goal_xml)
    schedule_ids = goal_data.schedule_ids or list(matches.keys())
    return ScheduleSnapshot(matches=matches, schedule_ids=schedule_ids, current_odds=goal_data.current_odds)


def filter_matches_by_date_window(
    matches: dict[str, MatchInfo],
    schedule_ids: Iterable[str],
    *,
    center_date: date | None = None,
    days_before: int = 7,
    days_after: int = 7,
) -> list[MatchInfo]:
    target_date = center_date or datetime.now().date()
    start_date = target_date - timedelta(days=max(0, days_before))
    end_date = target_date + timedelta(days=max(0, days_after))
    return [
        match
        for schedule_id in schedule_ids
        if (match := matches.get(schedule_id)) is not None and start_date <= match.match_time.date() <= end_date
    ]


def filter_today_matches(
    matches: dict[str, MatchInfo],
    schedule_ids: Iterable[str],
    *,
    today: date | None = None,
) -> list[MatchInfo]:
    return filter_matches_by_date_window(
        matches,
        schedule_ids,
        center_date=today,
        days_before=0,
        days_after=0,
    )


def read_selection_matches_from_workbook(
    output_path: str | Path,
    *,
    center_date: date | None = None,
    days_before: int = 7,
    days_after: int = 7,
) -> list[MatchInfo]:
    path = Path(output_path)
    if not path.exists():
        return []

    workbook = load_workbook(path, data_only=True)
    try:
        records = read_metadata_records(workbook)
    finally:
        workbook.close()

    target_date = center_date or datetime.now().date()
    matches: list[MatchInfo] = []
    seen: set[str] = set()
    for record in sorted(records, key=lambda item: (item.match_order, item.row_order, item.record_key)):
        if record.row_type != "match" or not record.schedule_id or record.schedule_id in seen:
            continue
        values = record.row_values
        match_time = infer_event_datetime_in_window(
            values[1] if len(values) > 1 else None,
            center_date=target_date,
            days_before=days_before,
            days_after=days_after,
        )
        if match_time is None:
            continue
        seen.add(record.schedule_id)
        if is_new_rank_column_row(values):
            home_team = cell_text(values, 4)
            away_team = cell_text(values, 7)
            status = cell_text(values, 2)
            score = cell_text(values, 5)
            total_line = cell_text(values, 19) if len(values) >= 22 else cell_text(values, 18)
        elif len(values) >= len(HEADERS_ROW):
            home_team = strip_rank_from_team(cell_text(values, 3))
            away_team = strip_rank_from_team(cell_text(values, 5))
            status = cell_text(values, 2)
            score = cell_text(values, 4)
            total_line = cell_text(values, 17)
        elif len(values) >= 19:
            home_team = strip_rank_from_team(cell_text(values, 3))
            away_team = strip_rank_from_team(cell_text(values, 5))
            status = cell_text(values, 2)
            score = cell_text(values, 4)
            total_line = cell_text(values, 16)
        else:
            home_team = cell_text(values, 2)
            away_team = cell_text(values, 4)
            status = ""
            score = cell_text(values, 3)
            total_line = cell_text(values, 6)
        matches.append(
            MatchInfo(
                schedule_id=record.schedule_id,
                league=cell_text(values, 0),
                event_time=cell_text(values, 1),
                home_team=home_team,
                score=score,
                away_team=away_team,
                half_score=cell_text(values, 5) if len(values) < len(HEADERS_ROW) else "",
                total_line=total_line,
                match_time=match_time,
                status=status,
            )
        )
    return matches


def infer_event_datetime_in_window(
    event_time: object,
    *,
    center_date: date,
    days_before: int = 7,
    days_after: int = 7,
) -> datetime | None:
    text = str(event_time or "")
    numbers = [int(value) for value in re.findall(r"\d{1,4}", text)]
    if len(numbers) < 3:
        return None

    start_date = center_date - timedelta(days=max(0, days_before))
    end_date = center_date + timedelta(days=max(0, days_after))
    candidates: list[datetime] = []

    def add_candidate(year: int, month: int, day: int, hour: int, minute: int) -> None:
        try:
            value = datetime(year, month, day, hour, minute)
        except ValueError:
            return
        if start_date <= value.date() <= end_date:
            candidates.append(value)

    if len(numbers) >= 5 and numbers[0] >= 1000:
        add_candidate(numbers[0], numbers[1], numbers[2], numbers[3], numbers[4])
    elif len(numbers) >= 4:
        month, day, hour, minute = numbers[-4], numbers[-3], numbers[-2], numbers[-1]
        for year in (center_date.year - 1, center_date.year, center_date.year + 1):
            add_candidate(year, month, day, hour, minute)
    else:
        day, hour, minute = numbers[-3], numbers[-2], numbers[-1]
        for offset in range(-max(0, days_before), max(0, days_after) + 1):
            candidate_date = center_date + timedelta(days=offset)
            if candidate_date.day == day:
                add_candidate(candidate_date.year, candidate_date.month, candidate_date.day, hour, minute)

    if not candidates:
        return None
    return min(candidates, key=lambda value: abs((value.date() - center_date).days))


def is_new_rank_column_row(values: tuple[object | None, ...]) -> bool:
    return len(values) > 7 and is_score_text(cell_text(values, 5))


def is_current_export_row(values: tuple[object | None, ...]) -> bool:
    if len(values) < len(HEADERS_ROW):
        return False
    old_total_time = values[16] if len(values) > 16 else None
    new_total_time = values[17] if len(values) > 17 else None
    if old_total_time and not new_total_time and not is_change_count_cell(old_total_time):
        return False
    return True


def is_change_count_cell(value: object | None) -> bool:
    if value in (None, ""):
        return True
    if isinstance(value, (int, float)):
        return True
    return re.fullmatch(r"\d+(?:\.0)?", str(value).strip()) is not None


def is_score_text(value: str) -> bool:
    return value == "-" or re.fullmatch(r"\d+\s*-\s*\d+", value) is not None


def cell_text(values: tuple[object | None, ...], index: int) -> str:
    if index >= len(values) or values[index] is None:
        return ""
    return str(values[index])


def parse_odds_history(
    html: str,
    *,
    kind: str,
    match_year: int,
    match_datetime: datetime | None = None,
    company_name: str = "澳*",
) -> list[OddsRecord]:
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table", id="oddsDetail")
    if table is None:
        return []

    rows = table.find_all("tr")
    if not rows:
        return []

    header_cells = rows[0].find_all(["th", "td"])
    headers = [cell.get_text(" ", strip=True) for cell in header_cells]
    try:
        company_index = headers.index(company_name)
    except ValueError:
        return []

    records: list[OddsRecord] = []
    for row in rows[1:]:
        cells = row.find_all("td")
        if len(cells) <= company_index:
            continue
        cell = cells[company_index]
        time_cell = cells[-1] if cells else None
        if time_cell is None:
            continue

        time_node = parse_history_time(
            time_cell.get_text(" ", strip=True),
            match_year=match_year,
            match_datetime=match_datetime,
        )
        parsed_cell = parse_odds_cell(cell)
        if time_node is None or parsed_cell is None:
            continue

        first, line, second = parsed_cell
        if kind == "asian":
            line = normalize_asian_handicap(line)
        else:
            line = normalize_total_line(line)
        if not line:
            continue
        records.append(OddsRecord(time_node, first, line, second))

    return sorted(records, key=lambda record: record.time_node)


def parse_change_detail_history(
    html: str,
    *,
    kind: str,
    match_year: int,
    match_datetime: datetime | None = None,
    status_filter: str = "即",
) -> list[OddsRecord]:
    soup = BeautifulSoup(html, "html.parser")
    records: list[OddsRecord] = []
    line_header = "盘" if kind == "asian" else "进球数"
    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        if not rows:
            continue
        headers = [cell.get_text(" ", strip=True) for cell in rows[0].find_all(["th", "td"])]
        if not {line_header, "变化时间", "状态"}.issubset(set(headers)):
            continue

        for row in rows[1:]:
            cells = row.find_all("td")
            if len(cells) < 7:
                continue
            status = cells[6].get_text(" ", strip=True)
            if status_filter and status != status_filter:
                continue
            time_node = parse_history_time(
                cells[5].get_text(" ", strip=True),
                match_year=match_year,
                match_datetime=match_datetime,
            )
            first = to_number(cells[2].get_text(" ", strip=True))
            line = cells[3].get_text(" ", strip=True)
            second = to_number(cells[4].get_text(" ", strip=True))
            if time_node is None or first is None or second is None:
                continue
            line = normalize_asian_handicap(line) if kind == "asian" else normalize_total_line(line)
            if not line:
                continue
            records.append(OddsRecord(time_node, first, line, second, status))
    unique = {
        (record.time_node, record.first, record.line, record.second, record.status): record
        for record in records
    }
    return sorted(unique.values(), key=lambda record: record.time_node)


def combine_export_rows(
    matches: dict[str, MatchInfo],
    schedule_ids: Iterable[str],
    asian_histories: dict[str, list[OddsRecord]],
    total_histories: dict[str, list[OddsRecord]],
    current_odds: dict[str, CurrentOdds] | None = None,
) -> list[list[object | None]]:
    return [
        list(record.row_values)
        for record in combine_export_records(matches, schedule_ids, asian_histories, total_histories, current_odds=current_odds)
    ]


def combine_export_records(
    matches: dict[str, MatchInfo],
    schedule_ids: Iterable[str],
    asian_histories: dict[str, list[OddsRecord]],
    total_histories: dict[str, list[OddsRecord]],
    current_odds: dict[str, CurrentOdds] | None = None,
) -> list[ExportRecord]:
    records: list[ExportRecord] = []
    current_odds = current_odds or {}
    for match_order, schedule_id in enumerate(schedule_ids):
        match = matches.get(schedule_id)
        if match is None:
            continue
        current = current_odds.get(schedule_id)
        asian_records = asian_histories.get(schedule_id, [])
        total_records = total_histories.get(schedule_id, [])
        row_count = max(len(asian_records), len(total_records), 1)

        for index in range(row_count):
            row: list[object | None] = [None] * len(HEADERS_ROW)
            row_type = "match" if index == 0 else "detail"
            if index == 0:
                row[:9] = [
                    match.league,
                    match.event_time,
                    match.status,
                    format_team_with_rank(match.home_team, match.home_rank),
                    match.score,
                    format_team_with_rank(match.away_team, match.away_rank),
                    current.asian_up if current else None,
                    current.asian_line if current else "",
                    current.asian_down if current else None,
                ]
            if index < len(asian_records):
                record = asian_records[index]
                row[9:15] = [
                    record.time_node,
                    record.first,
                    record.line,
                    record.second,
                    odds_change_count(asian_records, index),
                    handicap_change_count(asian_records, index),
                ]
            if index < len(total_records):
                record = total_records[index]
                row[15:21] = [
                    record.time_node,
                    record.first,
                    record.line,
                    record.second,
                    odds_change_count(total_records, index),
                    handicap_change_count(total_records, index),
                ]
            asian_key = odds_record_key(schedule_id, "asian", asian_records[index], index) if index < len(asian_records) else ""
            total_key = odds_record_key(schedule_id, "total", total_records[index], index) if index < len(total_records) else ""
            records.append(
                ExportRecord(
                    schedule_id=schedule_id,
                    record_key=f"{schedule_id}|row|{index}",
                    row_type=row_type,
                    match_order=match_order,
                    row_order=index,
                    row_values=tuple(row),
                    asian_key=asian_key,
                    total_key=total_key,
                )
            )
        records.append(
            ExportRecord(
                schedule_id=schedule_id,
                record_key=f"{schedule_id}|separator",
                row_type="separator",
                match_order=match_order,
                row_order=9999,
                row_values=tuple([None] * len(HEADERS_ROW)),
            )
        )
    return records


def record_signatures_by_schedule(records: Iterable[ExportRecord]) -> dict[str, tuple[tuple[object, ...], ...]]:
    grouped: dict[str, list[tuple[object, ...]]] = {}
    for record in records:
        grouped.setdefault(record.schedule_id, []).append(
            (
                record.record_key,
                record.row_type,
                record.row_order,
                record.row_values,
                record.asian_key,
                record.total_key,
            )
        )
    return {schedule_id: tuple(items) for schedule_id, items in grouped.items()}


def changed_schedule_ids(
    previous_signatures: dict[str, tuple[tuple[object, ...], ...]],
    current_records: Iterable[ExportRecord],
) -> list[str]:
    current_signatures = record_signatures_by_schedule(current_records)
    return [
        schedule_id
        for schedule_id, signature in current_signatures.items()
        if previous_signatures.get(schedule_id) != signature
    ]


def handicap_change_signatures_by_schedule(records: Iterable[ExportRecord]) -> dict[str, tuple[object, object]]:
    signatures: dict[str, tuple[object, object]] = {}
    for record in records:
        if record.row_type != "match":
            continue
        asian_count = record.row_values[14] if len(record.row_values) > 14 else None
        total_count = record.row_values[20] if len(record.row_values) > 20 else None
        signatures[record.schedule_id] = (asian_count, total_count)
    return signatures


def changed_handicap_schedule_ids(
    previous_signatures: dict[str, tuple[object, object]],
    current_records: Iterable[ExportRecord],
) -> list[str]:
    current_signatures = handicap_change_signatures_by_schedule(current_records)
    return [
        schedule_id
        for schedule_id, signature in current_signatures.items()
        if previous_signatures.get(schedule_id) != signature
    ]


def write_workbook(
    rows: list[list[object | None]],
    output_path: str | Path,
    *,
    template_path: str | Path | None = DEFAULT_TEMPLATE,
) -> Path:
    output_path = Path(output_path)
    workbook, sheet, row_styles = create_output_workbook(template_path)

    for row_index, row_values in enumerate(rows, start=2):
        style_source = 5 if all(value is None for value in row_values) else (2 if any(row_values[:9]) else 3)
        for column_index, value in enumerate(row_values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            apply_style(cell, row_styles.get(style_source, {}).get(column_index))
            if column_index in (10, 16):
                cell.number_format = "m/d/yy h:mm"
            apply_handicap_change_font(cell, column_index, value)

    workbook.save(output_path)
    return output_path


def sync_workbook(
    records: list[ExportRecord],
    output_path: str | Path,
    *,
    template_path: str | Path | None = DEFAULT_TEMPLATE,
) -> Path:
    output_path = Path(output_path)
    if output_path.exists():
        workbook = load_workbook(output_path)
        sheet = get_data_sheet(workbook)
        old_records = read_metadata_records(workbook)
        sheet = ensure_data_sheet_name(workbook, sheet)
        row_styles = apply_template_format(sheet, template_path) or capture_row_styles(sheet)
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row - 1)
        for column, header in enumerate(HEADERS_ROW, start=1):
            sheet.cell(1, column, header)
    else:
        workbook, sheet, row_styles = create_output_workbook(template_path)
        old_records = []

    merged_records = merge_export_records(old_records, records)
    write_records_to_sheet(sheet, merged_records, row_styles)
    write_metadata_sheet(workbook, merged_records)
    workbook.save(output_path)
    return output_path


def read_export_records_from_workbook(path: str | Path) -> list[ExportRecord]:
    path = Path(path)
    if not path.exists():
        return []
    workbook = load_workbook(path, data_only=True)
    try:
        return read_metadata_records(workbook)
    finally:
        workbook.close()


def capture_row_styles(sheet) -> dict[int, dict[int, object]]:
    row_styles: dict[int, dict[int, object]] = {}
    for row_number in (2, 3, 5):
        if sheet.max_row >= row_number:
            row_styles[row_number] = {
                column: copy.copy(sheet.cell(row_number, column)._style)
                for column in range(1, len(HEADERS_ROW) + 1)
            }
    return row_styles


def get_data_sheet(workbook):
    if DATA_SHEET_NAME in workbook.sheetnames:
        return workbook[DATA_SHEET_NAME]
    if LEGACY_DATA_SHEET_NAME in workbook.sheetnames:
        return workbook[LEGACY_DATA_SHEET_NAME]
    return workbook.active


def get_data_sheet_or_none(workbook):
    if DATA_SHEET_NAME in workbook.sheetnames:
        return workbook[DATA_SHEET_NAME]
    if LEGACY_DATA_SHEET_NAME in workbook.sheetnames:
        return workbook[LEGACY_DATA_SHEET_NAME]
    return None


def ensure_data_sheet_name(workbook, sheet):
    if sheet.title == DATA_SHEET_NAME:
        return sheet
    if DATA_SHEET_NAME not in workbook.sheetnames:
        sheet.title = DATA_SHEET_NAME
        return sheet
    return workbook[DATA_SHEET_NAME]


def apply_template_format(sheet, template_path: str | Path | None) -> dict[int, dict[int, object]]:
    template = resolve_runtime_path(template_path) if template_path else None
    if not template or not template.exists():
        return {}
    template_workbook = load_workbook(template)
    try:
        template_sheet = get_data_sheet(template_workbook)
        for column in range(1, len(HEADERS_ROW) + 1):
            letter = get_column_letter(column)
            sheet.column_dimensions[letter].width = template_sheet.column_dimensions[letter].width
        for row_number in (1, 2, 3, 5):
            for column in range(1, len(HEADERS_ROW) + 1):
                copy_cell_style(template_sheet.cell(row_number, column), sheet.cell(row_number, column))
        for row_number in (1, 2, 3, 5):
            sheet.row_dimensions[row_number].height = template_sheet.row_dimensions[row_number].height
        return capture_row_styles(sheet)
    finally:
        template_workbook.close()


def copy_cell_style(source, target) -> None:
    target.font = copy.copy(source.font)
    target.fill = copy.copy(source.fill)
    target.border = copy.copy(source.border)
    target.alignment = copy.copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy.copy(source.protection)


def merge_export_records(old_records: list[ExportRecord], new_records: list[ExportRecord]) -> list[ExportRecord]:
    old_match_order = {
        record.schedule_id: record.match_order
        for record in old_records
        if record.row_type == "match"
    }
    next_match_order = max(old_match_order.values(), default=-1) + 1
    new_match_order: dict[str, int] = {}
    aligned_new_records: list[ExportRecord] = []
    for record in new_records:
        if record.schedule_id in old_match_order:
            order = old_match_order[record.schedule_id]
        else:
            if record.schedule_id not in new_match_order:
                new_match_order[record.schedule_id] = next_match_order
                next_match_order += 1
            order = new_match_order[record.schedule_id]
        aligned_new_records.append(replace(record, match_order=order))

    merged: dict[str, ExportRecord] = {record.record_key: record for record in old_records}
    for record in aligned_new_records:
        old_record = merged.get(record.record_key)
        merged[record.record_key] = merge_export_record(old_record, record) if old_record else record
    return sorted(merged.values(), key=lambda record: (record.match_order, record.row_order, record.record_key))


def merge_export_record(old_record: ExportRecord, new_record: ExportRecord) -> ExportRecord:
    row_values = tuple(
        old_value if is_empty_cell(new_value) and not is_empty_cell(old_value) else new_value
        for old_value, new_value in zip(old_record.row_values, new_record.row_values)
    )
    return replace(
        new_record,
        row_values=row_values,
        asian_key=new_record.asian_key or old_record.asian_key,
        total_key=new_record.total_key or old_record.total_key,
    )


def is_empty_cell(value: object | None) -> bool:
    return value is None or value == ""


def write_records_to_sheet(sheet, records: list[ExportRecord], row_styles: dict[int, dict[int, object]]) -> None:
    for row_index, record in enumerate(records, start=2):
        style_source = {"match": 2, "detail": 3, "separator": 5}.get(record.row_type, 3)
        for column_index, value in enumerate(record.row_values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            apply_style(cell, row_styles.get(style_source, {}).get(column_index))
            if column_index in (10, 16):
                cell.number_format = "m/d/yy h:mm"
            apply_handicap_change_font(cell, column_index, value)


def read_metadata_records(workbook) -> list[ExportRecord]:
    if METADATA_SHEET not in workbook.sheetnames:
        return []
    visible = get_data_sheet_or_none(workbook)
    if visible is None:
        return []
    metadata = workbook[METADATA_SHEET]
    headers = [metadata.cell(1, column).value for column in range(1, metadata.max_column + 1)]
    required = set(METADATA_HEADERS)
    if not required.issubset(set(headers)):
        return []
    index = {name: headers.index(name) + 1 for name in METADATA_HEADERS}
    records: list[ExportRecord] = []
    for row in range(2, metadata.max_row + 1):
        visible_row = safe_int(metadata.cell(row, index["visible_row"]).value)
        if visible_row is None:
            continue
        row_values = tuple(visible.cell(visible_row, column).value for column in range(1, len(HEADERS_ROW) + 1))
        records.append(
            ExportRecord(
                schedule_id=str(metadata.cell(row, index["schedule_id"]).value or ""),
                record_key=str(metadata.cell(row, index["record_key"]).value or ""),
                row_type=str(metadata.cell(row, index["row_type"]).value or ""),
                match_order=safe_int(metadata.cell(row, index["match_order"]).value) or 0,
                row_order=safe_int(metadata.cell(row, index["row_order"]).value) or 0,
                row_values=row_values,
                asian_key=str(metadata.cell(row, index["asian_key"]).value or ""),
                total_key=str(metadata.cell(row, index["total_key"]).value or ""),
            )
        )
    return records


def write_metadata_sheet(workbook, records: list[ExportRecord]) -> None:
    if METADATA_SHEET in workbook.sheetnames:
        metadata = workbook[METADATA_SHEET]
        metadata.delete_rows(1, metadata.max_row)
    else:
        metadata = workbook.create_sheet(METADATA_SHEET)
    for column, header in enumerate(METADATA_HEADERS, start=1):
        metadata.cell(1, column, header)
    for row_index, record in enumerate(records, start=2):
        visible_row = row_index
        values = [
            METADATA_VERSION,
            visible_row,
            record.record_key,
            record.schedule_id,
            record.row_type,
            record.match_order,
            record.row_order,
            record.asian_key,
            record.total_key,
        ]
        for column, value in enumerate(values, start=1):
            metadata.cell(row_index, column, value)
    metadata.sheet_state = "hidden"


def crawl_complete_company_schedule(
    *,
    company_id: int = 1,
    limit: int | None = None,
    selected_schedule_ids: Iterable[str] | None = None,
    schedule_snapshot: ScheduleSnapshot | None = None,
    workers: int = 6,
    fetch_histories=None,
    progress_callback=None,
    cancel_event=None,
    log_to_console: bool = True,
) -> list[list[object | None]]:
    records = crawl_complete_company_schedule_records(
        company_id=company_id,
        limit=limit,
        selected_schedule_ids=selected_schedule_ids,
        schedule_snapshot=schedule_snapshot,
        workers=workers,
        fetch_histories=fetch_histories,
        progress_callback=progress_callback,
        cancel_event=cancel_event,
        log_to_console=log_to_console,
    )
    return [list(record.row_values) for record in records]


def crawl_complete_company_schedule_records(
    *,
    company_id: int = 1,
    limit: int | None = None,
    selected_schedule_ids: Iterable[str] | None = None,
    schedule_snapshot: ScheduleSnapshot | None = None,
    workers: int = 6,
    fetch_histories=None,
    progress_callback=None,
    cancel_event=None,
    log_to_console: bool = True,
    require_odds_records: bool = False,
) -> list[ExportRecord]:
    ensure_aomen_company_id(company_id)
    session = requests.Session()
    session.headers.update(HEADERS)
    schedule_snapshot = schedule_snapshot or fetch_company_schedule_snapshot(company_id=company_id, session=session)
    matches = schedule_snapshot.matches
    if selected_schedule_ids is not None:
        schedule_ids = [schedule_id for schedule_id in selected_schedule_ids if schedule_id]
    else:
        schedule_ids = schedule_snapshot.schedule_ids or list(matches.keys())
    if limit is not None:
        schedule_ids = schedule_ids[:limit]

    asian_histories: dict[str, list[OddsRecord]] = {}
    total_histories: dict[str, list[OddsRecord]] = {}
    total_count = len(schedule_ids)
    emit_progress(progress_callback, CrawlEvent("start", total=total_count))
    if log_to_console:
        print(f"准备抓取完整赛事 {total_count} 场")
    check_cancelled(cancel_event)
    fetch_histories = fetch_histories or fetch_match_histories

    with ThreadPoolExecutor(max_workers=max(1, workers)) as executor:
        futures = {}
        for schedule_id in schedule_ids:
            check_cancelled(cancel_event)
            futures[executor.submit(fetch_histories, session, schedule_id, matches.get(schedule_id))] = schedule_id
        finished = 0
        for future in as_completed(futures):
            check_cancelled(cancel_event)
            schedule_id = futures[future]
            try:
                asian_records, total_records = future.result()
                asian_histories[schedule_id] = asian_records
                total_histories[schedule_id] = total_records
            except Exception as exc:
                if log_to_console:
                    print(f"[失败] {schedule_id}: {exc}")
                emit_progress(progress_callback, CrawlEvent("error", total=total_count, completed=finished, schedule_id=schedule_id, message=str(exc)))
                asian_histories[schedule_id] = []
                total_histories[schedule_id] = []
            finished += 1
            emit_progress(progress_callback, CrawlEvent("match_done", total=total_count, completed=finished, schedule_id=schedule_id))
            if log_to_console and (finished % 5 == 0 or finished == total_count):
                print(f"已完成 {finished}/{total_count}")

    if require_odds_records:
        schedule_ids = [
            schedule_id
            for schedule_id in schedule_ids
            if asian_histories.get(schedule_id) or total_histories.get(schedule_id)
        ]

    records = combine_export_records(
        matches,
        schedule_ids,
        asian_histories,
        total_histories,
        current_odds=schedule_snapshot.current_odds,
    )
    emit_progress(progress_callback, CrawlEvent("complete", total=total_count, rows=len(records)))
    return records


def crawl_historical_date_records(
    match_date: date,
    *,
    company_id: int = 1,
    limit: int | None = None,
    workers: int = 6,
    schedule_snapshot: ScheduleSnapshot | None = None,
    fetch_histories=None,
    progress_callback=None,
    cancel_event=None,
    log_to_console: bool = True,
) -> list[ExportRecord]:
    ensure_aomen_company_id(company_id)
    session = requests.Session()
    session.headers.update(HEADERS)
    snapshot = schedule_snapshot or fetch_historical_schedule_snapshot(match_date, session=session)
    return crawl_complete_company_schedule_records(
        company_id=company_id,
        limit=limit,
        schedule_snapshot=snapshot,
        workers=workers,
        fetch_histories=fetch_histories,
        progress_callback=progress_callback,
        cancel_event=cancel_event,
        log_to_console=log_to_console,
        require_odds_records=True,
    )


def emit_progress(progress_callback, event: CrawlEvent) -> None:
    if progress_callback is not None:
        progress_callback(event)


def check_cancelled(cancel_event) -> None:
    if cancel_event is not None and cancel_event.is_set():
        raise CrawlCancelled("用户已取消抓取")


def fetch_match_histories(
    session: requests.Session,
    schedule_id: str,
    match: MatchInfo | None,
    *,
    retries: int = 2,
    retry_delay: float = 0.6,
    log_failures: bool = True,
) -> tuple[list[OddsRecord], list[OddsRecord]]:
    if match is None:
        return [], []

    asian_records = fetch_odds_records(
        session,
        change_detail_url(schedule_id, "handicap"),
        kind="asian",
        match=match,
        retries=retries,
        retry_delay=retry_delay,
        log_failures=log_failures,
    )
    total_records = fetch_odds_records(
        session,
        change_detail_url(schedule_id, "overunder"),
        kind="total",
        match=match,
        retries=retries,
        retry_delay=retry_delay,
        log_failures=log_failures,
    )
    return asian_records, total_records


def change_detail_url(schedule_id: str, page: str, company_id: int = AOMEN_COMPANY_ID) -> str:
    return f"{VIP_BASE_URL}/changeDetail/{page}.aspx?id={schedule_id}&companyID={company_id}&l=0"


def is_aomen_change_detail_url(url: str) -> bool:
    query = parse_qs(urlparse(url).query)
    company_values = query.get("companyID") or query.get("companyid") or []
    return str(AOMEN_COMPANY_ID) in {str(value) for value in company_values}


def is_aomen_change_detail_response(html: str) -> bool:
    decoded = unquote(unescape(html or ""))
    companies = re.findall(r"(?:[?&]|\b)company=([^&\"'<>\s]*)", decoded, flags=re.IGNORECASE)
    named_companies = [company for company in companies if company]
    if named_companies and any(company != AOMEN_COMPANY_NAME for company in named_companies):
        return False

    company_ids = re.findall(r"(?:[?&]|\b)companyid=(\d+)", decoded, flags=re.IGNORECASE)
    if company_ids and any(company_id != str(AOMEN_COMPANY_ID) for company_id in company_ids):
        return False

    return bool(named_companies or company_ids)


def ensure_aomen_company_id(company_id: int) -> None:
    if company_id != AOMEN_COMPANY_ID:
        raise ValueError(f"盘口公司必须是{AOMEN_COMPANY_NAME}(companyID={AOMEN_COMPANY_ID})")


def fetch_odds_records(
    session: requests.Session,
    url: str,
    *,
    kind: str,
    match: MatchInfo,
    retries: int,
    retry_delay: float,
    log_failures: bool,
) -> list[OddsRecord]:
    if not is_aomen_change_detail_url(url):
        if log_failures:
            print(f"[公司校验失败] {match.schedule_id} {kind}: {url}")
        return []
    try:
        html = fetch_text(session, url, referer=ENTRY_URL, retries=retries, retry_delay=retry_delay)
    except Exception as exc:
        if log_failures:
            print(f"[盘口失败] {match.schedule_id} {kind}: {exc}")
        return []
    if not is_aomen_change_detail_response(html):
        if log_failures:
            print(f"[公司校验失败] {match.schedule_id} {kind}: 返回页面不是{AOMEN_COMPANY_NAME}")
        return []
    return parse_change_detail_history(
        html,
        kind=kind,
        match_year=match.match_time.year,
        match_datetime=match.match_time,
    )


def fetch_text(
    session: requests.Session,
    url: str,
    *,
    referer: str,
    retries: int = 0,
    retry_delay: float = 0.6,
) -> str:
    last_error: Exception | None = None
    for attempt in range(retries + 1):
        try:
            response = session.get(url, headers={"Referer": referer}, timeout=30)
            response.raise_for_status()
            encoding = getattr(response, "apparent_encoding", None) or getattr(response, "encoding", None) or "utf-8"
            return response.content.decode(encoding, errors="replace")
        except Exception as exc:
            last_error = exc
            if attempt < retries and retry_delay > 0:
                time.sleep(retry_delay * (attempt + 1))
    assert last_error is not None
    raise last_error


def create_output_workbook(
    template_path: str | Path | None,
) -> tuple[Workbook, object, dict[int, dict[int, object]]]:
    row_styles: dict[int, dict[int, object]] = {}
    template = resolve_runtime_path(template_path) if template_path else None
    if template and template.exists():
        workbook = load_workbook(template)
        sheet = get_data_sheet(workbook)
        sheet = ensure_data_sheet_name(workbook, sheet)
        for row_number in (2, 3, 5):
            row_styles[row_number] = {
                column: copy.copy(sheet.cell(row_number, column)._style)
                for column in range(1, len(HEADERS_ROW) + 1)
            }
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row - 1)
        for column, header in enumerate(HEADERS_ROW, start=1):
            sheet.cell(1, column, header)
        return workbook, sheet, row_styles

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = DATA_SHEET_NAME
    widths = [9.25, 12.625, 8.625, 8.625, 16.75, 8.625, 8.625, 16.75, 9, 13, 9, 17.125, 11.875, 15.625, 11.875, 12, 12, 17.125, 11.875, 13, 11.875, 12, 12]
    for column, header in enumerate(HEADERS_ROW, start=1):
        cell = sheet.cell(1, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(column)].width = widths[column - 1]
    return workbook, sheet, row_styles


def resolve_runtime_path(path: str | Path) -> Path:
    runtime_path = Path(path)
    if runtime_path.is_absolute() or runtime_path.exists():
        return runtime_path
    bundle_dir = getattr(sys, "_MEIPASS", None)
    if bundle_dir:
        bundled_path = Path(bundle_dir) / runtime_path
        if bundled_path.exists():
            return bundled_path
    return runtime_path


def apply_style(cell, style) -> None:
    if style is not None:
        cell._style = copy.copy(style)
    if cell.alignment is None:
        cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_handicap_change_font(cell, column_index: int, value: object | None) -> None:
    if column_index not in (15, 21):
        return
    count = safe_int(value)
    if count is None:
        return
    font = copy.copy(cell.font)
    font.color = "FF0000" if count >= 4 else "1E3A8A"
    cell.font = font


def parse_odds_cell(cell) -> tuple[float | int, str, float | int] | None:
    text = cell.get_text(" ", strip=True)
    text = re.sub(r"\s+", " ", text).strip()
    matches = list(re.finditer(r"(?<![\w.])-?\d+(?:\.\d+)?", text))
    if len(matches) < 2:
        return None
    first_match, second_match = matches[-2], matches[-1]
    line = text[: first_match.start()].strip()
    first = to_number(first_match.group(0))
    second = to_number(second_match.group(0))
    if first is None or second is None:
        return None
    return first, line, second


def parse_history_time(
    text: str,
    *,
    match_year: int,
    match_datetime: datetime | None = None,
) -> datetime | None:
    match = re.search(r"(\d{1,2})-(\d{1,2})\s+(\d{1,2}):(\d{2})", text)
    if not match:
        return None
    month, day, hour, minute = map(int, match.groups())
    value = datetime(match_year, month, day, hour, minute)
    if match_datetime and value > match_datetime + timedelta(days=1):
        value = value.replace(year=value.year - 1)
    return value


def parse_match_datetime(fields: list[str]) -> datetime:
    raw = fields[12] if len(fields) > 12 else ""
    parts = [safe_int(part) for part in raw.split(",")]
    if len(parts) >= 6 and all(part is not None for part in parts[:6]):
        year, js_month, day, hour, minute, second = parts[:6]
        return datetime(year, js_month + 1, day, hour, minute, second)

    year = safe_int(fields[43] if len(fields) > 43 else "") or datetime.now().year
    month = day = 1
    date_match = re.search(r"(\d{1,2})-(\d{1,2})", fields[36] if len(fields) > 36 else "")
    if date_match:
        month, day = map(int, date_match.groups())
    hour = minute = 0
    time_match = re.search(r"(\d{1,2}):(\d{2})", fields[11] if len(fields) > 11 else "")
    if time_match:
        hour, minute = map(int, time_match.groups())
    return datetime(year, month, day, hour, minute)


def format_event_time(date_text: str, time_text: str) -> str:
    match = re.search(r"\d{1,2}-(\d{1,2})", date_text or "")
    if match:
        return f"{int(match.group(1))}日{time_text}"
    return time_text


def format_score(state: int | None, first: str, second: str) -> str:
    if state in (None, 0) or first == "" or second == "":
        return "-"
    return f"{first}-{second}"


def format_match_status(state: int | None) -> str:
    if state == -1:
        return "完"
    if state == 0:
        return "未"
    if state is None:
        return ""
    return "即"


def format_team_with_rank(team: str, rank: str) -> str:
    rank = (rank or "").strip()
    team = team or ""
    return f"[{rank}] {team}" if rank else team


def normalize_rank(value: str | None) -> str:
    matches = re.findall(r"\d+", value or "")
    return matches[-1] if matches else ""


def strip_rank_from_team(team: str) -> str:
    return re.sub(r"^\s*(?:\[\d+\]|\d+)\s+", "", team or "")


def asian_goal_to_text(value: str | float | int | None) -> str:
    number = safe_float(value)
    if number is None:
        return ""
    if abs(number) > 10:
        return f"{format_half_step(abs(number))}球" if number >= 0 else f"受让{format_half_step(abs(number))}球"
    index = int(round(abs(number) * 4))
    line = ASIAN_GOAL_NAMES[index] if index < len(ASIAN_GOAL_NAMES) else f"{format_half_step(abs(number))}球"
    return f"受让{line}" if number < 0 else line


def total_goal_to_text(value: str | float | int | None) -> str:
    number = safe_float(value)
    if number is None:
        return ""
    units = round(number * 4)
    whole = units // 4
    remainder = units % 4
    if remainder == 0:
        return str(whole)
    if remainder == 1:
        return f"{format_half_step(whole)}/{format_half_step(whole + 0.5)}"
    if remainder == 2:
        return format_half_step(whole + 0.5)
    return f"{format_half_step(whole + 0.5)}/{format_half_step(whole + 1)}"


def normalize_asian_handicap(line: str) -> str:
    value = re.sub(r"\s+", "", line or "")
    if value.startswith("*"):
        return "受让" + value.lstrip("*")
    return value


def normalize_total_line(line: str) -> str:
    return re.sub(r"\s+", "", line or "")


def odds_record_key(schedule_id: str, kind: str, record: OddsRecord, occurrence: int) -> str:
    # The visible workbook is row-oriented, but the hidden sheet keeps the
    # source odds identity so future sync logic can reason about record origin.
    time_key = record.time_node.strftime("%Y%m%d%H%M")
    return f"{schedule_id}|{kind}|{time_key}|{occurrence}"


def odds_change_count(records: list[OddsRecord], index: int) -> int | str:
    if index != 0:
        return ""
    return max(len(records) - 1, 0)


def handicap_change_count(records: list[OddsRecord], index: int) -> int | str:
    if index != 0:
        return ""
    lines = [record.line for record in records if record.line]
    return sum(1 for previous, current in zip(lines, lines[1:]) if previous != current)


def to_number(value: str | float | int | None) -> float | int | None:
    number = safe_float(value)
    if number is None:
        return None
    return int(number) if number.is_integer() else number


def safe_float(value: str | float | int | None) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    value = value.strip()
    if value == "":
        return None
    try:
        return float(value)
    except ValueError:
        return None


def safe_int(value: str | int | None) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def format_half_step(value: float | int) -> str:
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.1f}".rstrip("0").rstrip(".")


def strip_html(value: str) -> str:
    return unescape(re.sub(r"<[^>]+>", "", value or "")).strip()


def default_output_path() -> Path:
    return Path(SYNC_DEFAULT_OUTPUT)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="抓取 Titan007 完整赛事澳*盘口历史并导出 Excel")
    parser.add_argument("--company-id", type=int, default=1, help="盘口公司 ID，默认 1（澳*）")
    parser.add_argument("--template", default=DEFAULT_TEMPLATE, help="参考 Excel 模板路径")
    parser.add_argument("--output", default=None, help=f"输出 Excel 文件名，默认 {SYNC_DEFAULT_OUTPUT}，重复运行会同步同一文件")
    parser.add_argument("--history-date", default=None, help="抓取指定历史日期，格式 YYYYMMDD 或 YYYY-MM-DD")
    parser.add_argument("--limit", type=int, default=None, help="只抓取前 N 场，用于测试")
    parser.add_argument("--workers", type=int, default=6, help="盘口历史并发数")
    return parser.parse_args()


def main() -> None:
    # try:
    #     enforce_startup_time_limit()
    # except StartupBlocked as exc:
    #     raise SystemExit(str(exc)) from exc

    args = parse_args()
    if args.history_date:
        history_date = parse_cli_date(args.history_date)
        records = crawl_historical_date_records(
            history_date,
            company_id=args.company_id,
            limit=args.limit,
            workers=args.workers,
        )
        output_path = Path(args.output) if args.output else historical_output_path(Path.cwd(), history_date)
    else:
        records = crawl_complete_company_schedule_records(
            company_id=args.company_id,
            limit=args.limit,
            workers=args.workers,
        )
        output_path = Path(args.output) if args.output else default_output_path()
    sync_workbook(records, output_path, template_path=args.template)
    print(f"同步完成: {output_path.resolve()}，数据行 {len(records)}")


def parse_cli_date(value: str) -> date:
    text = (value or "").strip()
    for fmt in ("%Y%m%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise SystemExit("历史日期格式错误，请使用 YYYYMMDD 或 YYYY-MM-DD")


if __name__ == "__main__":
    main()
